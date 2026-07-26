"""Extract reproducible demographic inputs from the local ONE workbooks/CSVs."""

import csv
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "research" / "demografia-dominicana" / "data"
DATA.mkdir(parents=True, exist_ok=True)


def as_number(value):
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def extract_projection():
    source = ROOT / "posts" / "republica-en-un-grafico" / "2026-01-06-Perpectiva_del_desarrollo" / (
        "cuadro-estimaciones-proyecciones-población-total-por-año-según-región-provincia-2000-2030.xlsx"
    )
    book = load_workbook(source, read_only=True, data_only=True)
    sheet = book[book.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    years = []
    for col in range(1, len(rows[5])):
        value = rows[5][col]
        if isinstance(value, int) and 2000 <= value <= 2030:
            # The workbook places the year above the sex columns; the total
            # for that year is the preceding column.
            years.append((value, col - 1))

    output = []
    for row in rows[8:]:
        name = row[0] if row else None
        if not name:
            continue
        name = str(name).strip()
        level = "pais" if name == "Total país" else "region" if name.startswith("Región ") else "provincia"
        for year, total_col in years:
            if total_col + 2 >= len(row):
                continue
            total = as_number(row[total_col])
            male = as_number(row[total_col + 1])
            female = as_number(row[total_col + 2])
            if total is None:
                continue
            output.append({
                "location": name,
                "level": level,
                "year": year,
                "total": total,
                "male": male,
                "female": female,
            })

    path = DATA / "demografia_proyecciones_2000_2030.csv"
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=output[0].keys())
        writer.writeheader()
        writer.writerows(output)
    return path


def extract_age(source, year, age_column, weight_column):
    bins = [(0, 4), (5, 9), (10, 14), (15, 19), (20, 24), (25, 29), (30, 34),
            (35, 39), (40, 44), (45, 49), (50, 54), (55, 59), (60, 64),
            (65, 69), (70, 74), (75, 79), (80, 84), (85, 89), (90, 97)]
    totals = defaultdict(float)
    with source.open("r", encoding="cp1252", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            age = as_number(row.get(age_column))
            weight = as_number(row.get(weight_column))
            if age is None or weight is None or age < 0 or age > 97 or age == 99:
                continue
            bucket = next((f"{lo}-{hi}" for lo, hi in bins if lo <= age <= hi), None)
            if bucket:
                totals[bucket] += weight
    return [{"year": year, "age_band": band, "weighted_persons": totals[band]} for band in [f"{lo}-{hi}" for lo, hi in bins]]


def extract_age_distributions():
    sources = [
        (ROOT / "research" / "mercado-laboral-dominicano" / "data" / "raw" / "enhogar_2022" / "Personas_ENH22.csv", 2022, "P203", "F_ponderación"),
        (ROOT / "research" / "mercado-laboral-dominicano" / "data" / "raw" / "enhogar_2024" / "BD_ENH24_PERSONAS.csv", 2024, "P203", "FPONDERACION"),
    ]
    output = []
    for source, year, age_column, weight_column in sources:
        output.extend(extract_age(source, year, age_column, weight_column))
    path = DATA / "demografia_edad_enhogar_2022_2024.csv"
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=output[0].keys())
        writer.writeheader()
        writer.writerows(output)
    return path


if __name__ == "__main__":
    for path in (extract_projection(), extract_age_distributions()):
        print(path)
